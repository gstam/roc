import os
from cmath import exp
from openpyxl import Workbook, utils
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Protection, Font

from django.shortcuts import render, get_object_or_404, get_list_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import Http404, HttpResponse, FileResponse, HttpResponseRedirect


from django.views.generic import View
from django.views.generic.base import TemplateView
from django.views.generic.detail import DetailView
from django.views.generic.list import ListView
from django.views.generic.edit import CreateView, DeleteView, UpdateView, FormView

from django.db import transaction

from django.core.exceptions import *

from django.urls import reverse_lazy, reverse
from django.utils.translation import gettext_lazy as _

from django.conf import settings
from django.template.loader import get_template
from weasyprint import HTML
from django.contrib.staticfiles import finders
from django.template.loader import render_to_string
from config.settings import BASE_DIR

from .models import *
from .forms import AccountForm, CreditForm, DebitForm, RecallForm, RequestForPaymentOrInvoiceForm, PaymentForm, RecallRequestToPdfForm
from .utils import *
from decimal import *
import calendar
from .data import *

# Create your views here.
class ExpenditureRegisterCreateView(LoginRequiredMixin, CreateView):
    model = ExpenditureRegister
    fields = '__all__' 


class ExpenditureRegisterUpdateView(LoginRequiredMixin, UpdateView):
    model = ExpenditureRegister
    fields = '__all__' 


class ExpenditureRegisterDetailView(LoginRequiredMixin, DetailView):
    model = ExpenditureRegister


class ExpenditureRegisterListView(LoginRequiredMixin, ListView):
    model = ExpenditureRegister

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.order_by('-year')


class ExpenditureRegisterDeleteView(LoginRequiredMixin, DeleteView):
    model = ExpenditureRegister
    # template_name = 'expenditureregister_confirm_delete.html'
    success_url = reverse_lazy('register_list')


months = [
    ('1','Ιανουάριος'),
    ('2','Φεβρουάριος'),
    ('3','Μάρτιος'),
    ('4','Απρίλιος'),
    ('5','Μάιος'),
    ('6','Ιούνιος'),
    ('7','Ιούλιος'),
    ('8','Αύγουστος'),
    ('9','Σεπτέμβριος'),
    ('10','Οκτώβριος'),
    ('11','Νοέμβριος'),
    ('12','Δεκέμβριος'),
    ]

class RegisterReportView(LoginRequiredMixin, ListView):
    model = Account
    template_name = 'expenditure_register/report.html'

    def get_queryset(self, ):
        queryset = super().get_queryset()
        queryset = Account.objects.filter(expenditure_register = self.kwargs['register_pk'])
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_month = datetime.datetime.today().month
        current_year =  datetime.datetime.today().year
        reference_year = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk']).year
        reference_month = self.request.GET.get('month_selection','')
        if reference_month:
            reference_month = int(reference_month)
        else:
            if current_month == 1:
                reference_month = current_month 
            else:
                reference_month = current_month - 1

        if reference_year < current_year:
            valid_months = months
        else:
            valid_months = months[:current_month]
        
        reference_day = calendar.monthrange(reference_year, int(reference_month))[1]
        reference_date = datetime.date(year=reference_year, month=reference_month, day=reference_day)
        report_data = compute_report_data(self.kwargs['register_pk'], reference_date) 
        context['valid_months'] = valid_months
        context['reference_year'] = str(reference_year)
        context['reference_month'] = str(reference_month)
        context['reference_day'] = str(reference_day)
        context['data_list'] = report_data
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account_selection'] = [account_category[0] for account_category in account_CATEGORIES_TYPE_CHOICES]
        return context


class RegisterReportToXlsxView(LoginRequiredMixin, View):
    
    def get(self, request, *args, **kwargs):
        reference_date = datetime.date(year=self.kwargs['reference_year'], 
                                       month=self.kwargs['reference_month'], 
                                       day=self.kwargs['reference_day'])
        report_data = compute_report_data(self.kwargs['register_pk'], reference_date)
        wb = Workbook()
        ws = wb.active
        # Formatting rows, columns and cells of the spreadsheet.
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        ws.append(('Έτος (4)','Περίοδος (3)','Κατηγορία Π/Υ (3)','Φορέας (2)','Ειδικός Φορέας','ΑΛΕ (4)','(4) Εγκεκριμένη Πίστωση Π/Υ (5)','(5) Αναμόρφωση Π/Υ (+/-)','(6) Διαμόρφωση Π/Υ = (4)+(5)','(7) Ποσοστό Διάθεσης Πιστώσεων','(11) Ανειλλημμένη Δέσμευση (Ποσό Δέσμευσης)','(12) Υπολειπόμενη προς Διάθεση πίστωση (12)=(6)*(7)-(11)','(22) Ποσό τιμολογίου ή άλλου ισοδύναμου εγγράφου','(36) Συνολικό Ποσό πληρωμής','(37) Συνολικό ποσό πληρωμής Ληξ. Οφειλ > 90 προς Τρίτους','(38) Εκ του οποίου συμψηφ./παρακρ. Υπέρ του Δημοσ. ή ΟΚΑ','(39) Εκκρεμείς Δεσμεύσεις (39)=(11)-(36)','(40) Σύνολο Απλήρωτων Υποχρεώσεων (40)=(22)-(36)','(41) Απλήρωτες Υποχρεώσεις προς Γενική Κυβέρνηση','(42) Απλήρωτες Υποχρεώσεις προς Τρίτους','(43) Εκκρεμείς Οφειλές προς Γενική Κυβέρνηση 1 - 30 ημ','(44) Εκκρεμείς Οφειλές προς Τρίτους 1 - 30 ημ','(45) Εκκρεμείς Οφειλές προς Γενική Κυβέρνηση 31 - 60 ημ','(46) Εκκρεμείς Οφειλές προς Τρίτους 31 - 60 ημ','(47) Εκκρεμείς Οφειλές προς Γενική Κυβέρνηση 61 - 90 ημ','(48) Εκκρεμείς Οφειλές προς Τρίτους 61 - 90 ημ','(49) Ληξιπρόθεσμες Οφειλές προς Γεν. Κυβέρνηση> 90 ημ','(50) Ληξιπρόθεσμες Οφειλές προς Τρίτους > 90 ημ','(51) Ληξιπρόθεσμες Οφειλές προς Τρίτους μη εγχ. Πιστωτές > 90 ημ'))
        ws.row_dimensions[1].height = 65
        for col in range(1,30):
            column_letter = utils.cell.get_column_letter(col)
            ws.column_dimensions[column_letter].width = 25
            c = ws.cell(column=col, row=1)
            c.alignment = Alignment(horizontal='general', vertical='bottom', wrap_text=True)
            c.font = Font(bold=True)
            c.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),  top=Side(border_style='thin'),  bottom=Side(border_style='thin'))
            if col <= 6:
                c.fill = PatternFill(fill_type = "solid", start_color="f1c40f")
            elif col >=7 and col <=10:
                c.fill = PatternFill(fill_type = "solid", start_color="d5f5e3")
            elif col >=11 and col <= 16:
                c.fill = PatternFill(fill_type = "solid", start_color="85c1e9")
            elif col >= 17 and col <= 20:
                c.fill = PatternFill(fill_type = "solid", start_color="f9e79f")
            elif col >= 21 and col <= 26:
                c.fill = PatternFill(fill_type = "solid", start_color="f5b041")
            elif col >= 27 and col <= 30:
                c.fill = PatternFill(fill_type = "solid", start_color="af7ac5")
            else:
                pass
        
        # Insert data into spreadsheet
        row_count = 1
        for account, initial_credit, reformed_credit, total_credit, disposed_percentage_list, total_debit, total_residual_credit_list, total_invoice, total_payment, total_90days_delayed_payment_to_third_parties, total_90days_delayed_setoff_payment_to_third_parties, total_pending_debits, total_pending_payments, invoice_amount_owned_to_general_goverment, total_invoice_amount_owned_to_third_parties, owned_to_goverment_1_to_30_days, owned_to_third_party_1_to_30_days, owned_to_goverment_31_to_60_days, owned_to_third_party_31_to_60_days, owned_to_goverment_61_to_90_days, owned_to_third_party_61_to_90_days, owned_to_goverment_91_to_end_days, owned_to_third_party_91_to_end_days, owned_to_third_party_from_abroad_91_to_end_days in report_data:
            row = (str(self.kwargs['reference_year']), self.kwargs['reference_month'], 'E99', '1019', '206-9922000', str(account.number), initial_credit, reformed_credit, total_credit, disposed_percentage_list, total_debit, total_residual_credit_list, total_invoice, total_payment, total_90days_delayed_payment_to_third_parties, total_90days_delayed_setoff_payment_to_third_parties, total_pending_debits, total_pending_payments, invoice_amount_owned_to_general_goverment, total_invoice_amount_owned_to_third_parties, owned_to_goverment_1_to_30_days, owned_to_third_party_1_to_30_days, owned_to_goverment_31_to_60_days, owned_to_third_party_31_to_60_days, owned_to_goverment_61_to_90_days, owned_to_third_party_61_to_90_days, owned_to_goverment_91_to_end_days, owned_to_third_party_91_to_end_days, owned_to_third_party_from_abroad_91_to_end_days)
            ws.append(row)
            row_count += 1

        # More formatting... adding border lines to cells with data.
        for row in ws.iter_rows(min_row=1, min_col=1, max_row=row_count, max_col=29):
            for cell in row:
                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),  top=Side(border_style='thin'),  bottom=Side(border_style='thin'))
        wb.save('./expenditure_register/templates/expenditure_register/report.xlsx')
        return FileResponse(open('./expenditure_register/templates/expenditure_register/report.xlsx', 'rb'), as_attachment=True, filename='Report.xlsx')


# def link_callback(uri, rel):
#     """
#     Convert HTML URIs to absolute system paths so xhtml2pdf can access those
#     resources
#     """
#     result = finders.find(uri)
#     if result:
#         if not isinstance(result, (list, tuple)):
#             result = [result]
#         result = list(os.path.realpath(path) for path in result)
#         path = result[0]
#     else:
#         sUrl = settings.STATIC_URL        # Typically /static/
#         sRoot = settings.STATIC_ROOT      # Typically /home/userX/project_static/
#         mUrl = settings.MEDIA_URL         # Typically /media/
#         mRoot = settings.MEDIA_ROOT       # Typically /home/userX/project_static/media/

#         if uri.startswith(mUrl):
#             path = os.path.join(mRoot, uri.replace(mUrl, ""))
#         elif uri.startswith(sUrl):
#             path = os.path.join(sRoot, uri.replace(sUrl, ""))
#         else:
#             return uri

#     # make sure that file exists
#     if not os.path.isfile(path):
#             raise Exception(
#                 'media URI must start with %s or %s' % (sUrl, mUrl)
#             )
#     return path

class GenerateRecallRequestFormView(LoginRequiredMixin, FormView):
    template_name = os.path.join(BASE_DIR, 'expenditure_register/templates/expenditure_register/recall_request_pdf_form.html')
    form_class = RecallRequestToPdfForm
    success_url = 'recall_request_print_pdf'

    def post(self, request):
        template_path = os.path.join(BASE_DIR, 'expenditure_register/templates/expenditure_register/recall_request_pdf_document.html')
        recipients = self.request.POST.get('recipients','').split(",")
        amount_in_text = reference_month = self.request.POST.get('amount_in_text','')
        amount = self.request.POST.get('amount','')
        protocol_number = self.request.POST.get('protocol_number','')
        protocol_date = self.request.POST.get('protocol_date','')
        year = self.request.POST.get('year','')
        account = self.request.POST.get('account','')
        # 'data': data,
        context = {'recipients': recipients, 'amount_in_text': amount_in_text, 'amount': amount, 'protocol_number': protocol_number, 'protocol_date': protocol_date, 'year': year, 'account': account, 'charset': 'iso-8859-7'}
        # # Create a Django response object, and specify content_type as pdf
        # response = HttpResponse(content_type='application/pdf')
        # response['Content-Disposition'] = 'attachment; filename="report.pdf"'
        # # find the template and render it.
        # template = get_template(template_path)
        # source_html_file = template.render(context)#.encode('iso-8859-7')
        base_url = os.path.dirname(os.path.realpath(__file__))
        content_string = render_to_string(template_path, context).encode('iso-8859-7')
        
        # with open(os.path.join(BASE_DIR, 'expenditure_register/templates/expenditure_register/sample_to_send.html'), 'w') as static_file:
        #     static_file.write(content)
        #HTML(filename = '/roc/expenditure_register/templates/expenditure_register/simple.html').write_pdf('/tmp/report.pdf')
        HTML(string=content_string, base_url=base_url).write_pdf('/tmp/report.pdf')
        return  FileResponse(open('/tmp/report.pdf', 'rb'), as_attachment=True, filename='report.pdf')
        

# class GenerateRecallRequestInPdfView(LoginRequiredMixin, View):

#     def get(self, request, *args, **kwargs):
#         template_path = os.path.join(BASE_DIR, 'expenditure_register/templates/expenditure_register/recall_request_pdf_document.html')
#         recipients = self.request.GET.get('recipients','').split(",")
#         amount_in_text = reference_month = self.request.GET.get('amount_in_text','')
#         amount = self.request.GET.get('amount','')
#         protocol_number = self.request.GET.get('protocol_number','')
#         protocol_date = self.request.GET.get('protocol_date','')
#         year = self.request.GET.get('year','')
#         account = self.request.GET.get('account','')
#         # 'data': data,
#         context = {'recipients': recipients, 
#                    'amount_in_text': amount_in_text, 
#                    'amount': amount, 
#                    'protocol_number': protocol_number, 
#                    'protocol_date': protocol_date, 
#                    'year': year, 
#                    'account': account, 
#                    'charset': 'iso-8859-7'}
#         # # Create a Django response object, and specify content_type as pdf
#         # response = HttpResponse(content_type='application/pdf')
#         # response['Content-Disposition'] = 'attachment; filename="report.pdf"'
#         # # find the template and render it.
#         # template = get_template(template_path)
#         # source_html_file = template.render(context)#.encode('iso-8859-7')
#         base_url = os.path.dirname(os.path.realpath(__file__))
#         content_string = render_to_string(template_path, context).encode('iso-8859-7')
#         # with open(os.path.join(BASE_DIR, 'expenditure_register/templates/expenditure_register/sample_to_send.html'), 'w') as static_file:
#         #     static_file.write(content)
#         HTML(string=content_string, base_url=base_url).write_pdf('/tmp/weasyprint-website.pdf')
#         return  FileResponse(open('/tmp/weasyprint-website.pdf', 'rb'), as_attachment=True, filename='Report.pdf')

 
class AccountCreateView(LoginRequiredMixin, CreateView):
    model = Account
    # fields = '__all__' 
    form_class = AccountForm

    
    def get_initial(self):
        expenditure_register = get_object_or_404(ExpenditureRegister, id=self.kwargs['register_pk'])
        return {'expenditure_register': expenditure_register,}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('account_list', args = (self.kwargs['register_pk'],))


class AccountUpdateView(LoginRequiredMixin, UpdateView):
    model = Account
    fields = '__all__' 

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('account_list', args = (self.kwargs['register_pk'],))


class AccountDetailView(LoginRequiredMixin, DetailView):
    model = Account


class AccountListView(LoginRequiredMixin, ListView):
    model = Account
    
    # Get the set of exit accounts for a specific expenditure register.
    def get_queryset(self, ):
        queryset = super().get_queryset()
        account_subcategory = self.request.GET.get('account_selection','')
        if account_subcategory:
            self.request.session['account_subcategory'] = account_subcategory

        # Check if key: 'account_subcategory' has been set in the session dictionary.
        if 'account_subcategory' in self.request.session:
            if self.request.session['account_subcategory'] and self.request.session['account_subcategory'] != 'ΟΛΟΙ':
                queryset = Account.objects.filter(expenditure_register = self.kwargs['register_pk']).filter(subcategory = self.request.session['account_subcategory'])
            else:
                queryset = Account.objects.filter(expenditure_register = self.kwargs['register_pk'])
        else:
            queryset = Account.objects.filter(expenditure_register = self.kwargs['register_pk'])
        return queryset

    # Add to context the data of the Expediture Register so that the user can see the register book it works on.
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # balance_list = [get_account_balance(account.pk) for account in context['object_list']]
        total_credit_list = [compute_total_credit(account.pk) for account in context['object_list']]
        total_debit_list = [compute_total_debit(account.pk) for account in context['object_list']]
        total_residual_credit_list = [c - d for c, d in zip(total_credit_list, total_debit_list)] 
        total_recall_list = [compute_total_recall(account.pk) for account in context['object_list']]
        total_invoice_list = [compute_total_invoice_amount(account.pk) for account in context['object_list']]
        total_amount_to_recall = [c - i - r for c, i, r in zip(total_credit_list, total_invoice_list, total_recall_list)] 
        total_payment_list = [compute_total_payment_amount(account.pk) for account in context['object_list']]
        total_pending_debits = [d - p for d, p in zip(total_debit_list, total_payment_list)]
        total_pending_payments = [i - p for i, p in zip(total_invoice_list, total_payment_list)]

        context['data_list'] = zip(list(context['object_list']), total_credit_list, total_debit_list, total_residual_credit_list, total_recall_list, total_invoice_list, total_amount_to_recall, total_payment_list, total_pending_debits, total_pending_payments)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account_selection'] = [account_category[0] for account_category in account_CATEGORIES_TYPE_CHOICES]
        context['selected_account_type'] = self.request.session.get('account_subcategory', 'ΟΛΟΙ')
        return context


class AccountDeleteView(LoginRequiredMixin, DeleteView):
    model = Account
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        return context
  
    # template_name = 'expenditureregister_confirm_delete.html'
    def get_success_url(self) -> str:
        return reverse_lazy('account_list', args = (self.kwargs['register_pk'],))


class CreditCreateView(LoginRequiredMixin, CreateView):
    model = Credit
    form_class = CreditForm

    def get_initial(self):
        initial = super().get_initial()
        account = Account.objects.get(pk = self.kwargs['account_pk'])
        initial['account'] = account
        initial['initial_credit'] = Decimal(0.0) #Credit.objects.get(pk = self.kwargs['pk']).credit
        initial['total_credit'] = compute_total_credit(self.kwargs['account_pk'])
        initial['total_debit'] = compute_total_debit(self.kwargs['account_pk'])
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     new_credit = form.save(commit = False)
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     # account.balance += new_credit.credit * new_credit.disposed_percentage/Decimal(100.00)
    #     account.save()
    #     form.save()
    #     return super().form_valid(form)
            
    def get_success_url(self):
        return reverse_lazy('credit_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class CreditUpdateView(LoginRequiredMixin, UpdateView):
    model = Credit
    # fields = '__all__' 
    form_class = CreditForm

    def get_initial(self):
        initial = super().get_initial()
        account = Account.objects.get(pk = self.kwargs['account_pk'])
        initial['account'] = account
        initial['initial_credit'] = Credit.objects.get(pk = self.kwargs['pk']).credit
        initial['total_credit'] = compute_total_credit(self.kwargs['account_pk'])
        initial['total_debit'] = compute_total_debit(self.kwargs['account_pk'])
        return initial


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     old_credit = Credit.objects.get(pk = self.kwargs['pk'])
    #     new_credit = form.save(commit = False)
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     account.balance =  account.balance + new_credit.credit * new_credit.disposed_percentage/Decimal(100.00) - old_credit.credit * old_credit.disposed_percentage/Decimal(100.00)
    #     account.save()
    #     form.save()
    #     return super().form_valid(form)

    def get_success_url(self) -> str:
        return reverse_lazy('credit_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class CreditDetailView(LoginRequiredMixin, DetailView):
    model = Credit


class CreditListView(LoginRequiredMixin, ListView):
    model = Credit
    
    # Get the set of exit accounts for a specific expenditure register.
    def get_queryset(self):
        # queryset = super().get_queryset()
        queryset = Credit.objects.filter(account = self.kwargs['account_pk'])
        return queryset

    # Add to context the data of the Account so that the user can see the register book it works on.
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        return context


class CreditDeleteView(LoginRequiredMixin, DeleteView):
    model = Credit

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        credit_to_delete = Credit.objects.get(pk = self.kwargs['pk']).credit
        total_credit = compute_total_credit(self.kwargs['account_pk'])
        total_debit = compute_total_debit(self.kwargs['account_pk'])
        if total_credit - credit_to_delete < total_debit:
            context['not_ok_to_delete'] = 1
        else:
            context['ok_to_delete'] = 1
        return context
  
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     credit = Credit.objects.get(pk = self.kwargs['pk'])
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     account.balance = account.balance - credit.credit*credit.disposed_percentage/Decimal(100.00)
    #     account.save()
    #     return super().form_valid(form)

    # template_name = 'expenditureregister_confirm_delete.html'
    def get_success_url(self) -> str:
        return reverse_lazy('credit_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class DebitCreateView(LoginRequiredMixin, CreateView):
    model = Debit
    form_class = DebitForm

    def get_initial(self):
        initial = super().get_initial()
        account = get_object_or_404(Account, pk=self.kwargs['account_pk'])
        initial['account'] = account
        initial['total_credit'] = compute_total_credit(self.kwargs['account_pk'])
        initial['total_debit'] = compute_total_debit(self.kwargs['account_pk'])
        initial['initial_debit_value'] = Decimal(0.00)
        initial['invoices_charged_on_debit'] = Decimal(0.00)

        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        # context['balance'] = get_account_balance(self.kwargs['account_pk'])
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     new_debit = form.save(commit = False)
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     account.balance -= new_debit.debit
    #     account.save()
    #     form.save()
    #     return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('debit_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class DebitUpdateView(LoginRequiredMixin, UpdateView):
    model = Debit
    # fields = '__all__' 
    form_class = DebitForm

    def get_initial(self):
        initial = super().get_initial()
        account = get_object_or_404(Account, pk=self.kwargs['account_pk'])
        initial['account'] = account
        initial['total_credit'] = compute_total_credit(self.kwargs['account_pk'])
        initial['total_debit'] = compute_total_debit(self.kwargs['account_pk'])
        initial['initial_debit_value'] = get_object_or_404(Debit, pk = self.kwargs['pk']).debit
        initial['invoices_charged_on_debit'] = compute_total_invoice_amount_charged_on_debit(self.kwargs['account_pk'], self.kwargs['pk'])
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     old_debit = Debit.objects.get(pk = self.kwargs['pk'])
    #     new_debit = form.save(commit = False)
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     account.balance = account.balance + old_debit.debit - new_debit.debit
    #     account.save()
    #     form.save()
    #     return super().form_valid(form)

    def get_success_url(self) -> str:
        return reverse_lazy('debit_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class DebitDetailView(LoginRequiredMixin, DetailView):
    model = Debit


class DebitListView(LoginRequiredMixin, ListView):
    model = Debit
    
    # Get the set of exit accounts for a specific expenditure register.
    def get_queryset(self):
        # queryset = super().get_queryset()
        queryset = Debit.objects.filter(account = self.kwargs['account_pk'])
        return queryset

    # Add to context the data of the Account so that the user can see the register book it works on.
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        return context


class DebitDeleteView(LoginRequiredMixin, DeleteView):
    model = Debit

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        total_invoice_amount = compute_total_invoice_amount_charged_on_debit(self.kwargs['account_pk'], self.kwargs['pk'])
        if total_invoice_amount > 0:
            context['not_ok_to_delete'] = 1
        else:
            context['ok_to_delete'] = 1
            # raise ValidationError(_(f'Δεν μπορείτε να διαγράψετε ανάληψη υποχρέωσης επί της οποίας έχουν εκδοθεί καταστάσεις πληρωμής ή τιμολόγια.'))
        return context
  
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     old_debit = Debit.objects.get(pk = self.kwargs['pk'])
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     account.balance += old_debit.debit
    #     account.save()
    #     return super().form_valid(form)


    # template_name = 'expenditureregister_confirm_delete.html'
    def get_success_url(self) -> str:
        return reverse_lazy('debit_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class RecallCreateView(LoginRequiredMixin, CreateView):
    model = Recall
    form_class = RecallForm

    def get_initial(self):
        initial = super().get_initial()
        account = get_object_or_404(Account, pk=self.kwargs['account_pk'])
        initial['account'] = account
        initial['account_pk'] = self.kwargs['account_pk']
        # initial['balance'] = get_account_balance(self.kwargs['account_pk'])
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        # context['balance'] = get_account_balance(self.kwargs['account_pk'])
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     new_debit = form.save(commit = False)
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     # account.balance -= new_debit.debit
    #     account.save()
    #     form.save()
    #     return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy('recall_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class RecallUpdateView(LoginRequiredMixin, UpdateView):
    model = Recall
    # fields = '__all__' 
    form_class = RecallForm

    def get_initial(self):
        initial = super().get_initial()
        account = get_object_or_404(Account, pk=self.kwargs['account_pk'])
        initial['account'] = account
        initial['account_pk'] = self.kwargs['account_pk']
        # print((self.fields))
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     old_debit = Debit.objects.get(pk = self.kwargs['pk'])
    #     new_debit = form.save(commit = False)
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     # account.balance = account.balance + old_debit.debit - new_debit.debit
    #     # account.save()
    #     form.save()
    #     return super().form_valid(form)

    def get_success_url(self) -> str:
        return reverse_lazy('recall_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class RecallDetailView(LoginRequiredMixin, DetailView):
    model = Recall


class RecallListView(LoginRequiredMixin, ListView):
    model = Recall
    
    # Get the set of exit accounts for a specific expenditure register.
    def get_queryset(self):
        # queryset = super().get_queryset()
        queryset = Recall.objects.filter(account = self.kwargs['account_pk'])
        return queryset

    # Add to context the data of the Account so that the user can see the register book it works on.
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        return context


class RecallDeleteView(LoginRequiredMixin, DeleteView):
    model = Recall

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        return context
  
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     old_debit = Debit.objects.get(pk = self.kwargs['pk'])
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     # account.balance += old_debit.debit
    #     account.save()
    #     return super().form_valid(form)


    # template_name = 'expenditureregister_confirm_delete.html'
    def get_success_url(self) -> str:
        return reverse_lazy('recall_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],))


class RequestForPaymentOrInvoiceCreateView(LoginRequiredMixin, CreateView):
    model = RequestForPaymentOrInvoice
    form_class = RequestForPaymentOrInvoiceForm

    def get_initial(self):
        # account = get_object_or_404(Account, pk=self.kwargs['account_pk'])
        initial = super().get_initial()
        initial['debit'] = get_object_or_404(Debit, pk=self.kwargs['debit_pk'])
        initial['invoice_charge_amount'] = Decimal(0.0) #get_object_or_404(RequestForPaymentOrInvoice, pk=self.kwargs['pk']).invoice_charge_amount
        initial['invoices_charged_on_debit'] = compute_total_invoice_amount_charged_on_debit(self.kwargs['account_pk'], self.kwargs['debit_pk'])
        initial['invoice_payment'] = Decimal(0.0)
        return initial #'account': account, 

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit']  = Debit.objects.get( pk=self.kwargs['debit_pk'])
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     new_invoice = form.save(commit = False)
    #     debit = Debit.objects.get(pk = self.kwargs['debit_pk'])
    #     invoices_query_set = RequestForPaymentOrInvoice.objects.filter(debit = self.kwargs['debit_pk'])
    #     s = Decimal(0.0)
    #     for q in invoices_query_set:
    #         s +=(q.invoice_charge_amount)
    #     s += new_invoice.invoice_charge_amount
    #     print(s)
    #     return super().form_valid(form)

        # account.balance += new_credit.credit * new_credit.disposed_percentage/Decimal(100.00)
        # account.save()
        
            
    def get_success_url(self):
        return reverse_lazy('invoice_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],self.kwargs['debit_pk']))


class RequestForPaymentOrInvoiceUpdateView(LoginRequiredMixin, UpdateView):
    model = RequestForPaymentOrInvoice
    # fields = '__all__' 
    # template_name = "requestforpaymentorinvoice_form.html"
    form_class = RequestForPaymentOrInvoiceForm
    
    def get_initial(self):
        initial = super().get_initial()
        initial['debit'] = get_object_or_404(Debit, pk=self.kwargs['debit_pk'])
        # initial['invoice_charge_amount'] = Decimal(0.0) #get_object_or_404(RequestForPaymentOrInvoice, pk=self.kwargs['pk']).invoice_charge_amount
        initial['invoices_charged_on_debit'] = compute_total_invoice_amount_charged_on_debit(self.kwargs['account_pk'], self.kwargs['debit_pk'])
        try:
            invoice_payment = Payment.objects.get(invoice = self.kwargs['pk']).payment_amount
        except Payment.DoesNotExist: #OneToOne relationship so 
            invoice_payment = Decimal(0.0)
        initial['invoice_payment'] = invoice_payment
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit'] = Debit.objects.get( pk=self.kwargs['debit_pk'])
        try:
            invoice_payment = Payment.objects.get(invoice = self.kwargs['pk']).payment_amount
        except Payment.DoesNotExist:
            invoice_payment = Decimal(0.0)
        context['invoice_payment'] = invoice_payment
        
        return context
    
    def get_success_url(self) -> str:
        return reverse_lazy('invoice_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'], self.kwargs['debit_pk']))


class RequestForPaymentOrInvoiceDetailView(LoginRequiredMixin, DetailView):
    model = RequestForPaymentOrInvoice

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit'] = Debit.objects.get( pk=self.kwargs['debit_pk'])
        return context


class RequestForPaymentOrInvoiceListView(LoginRequiredMixin, ListView):
    model = RequestForPaymentOrInvoice
    
    # Get the set of exit accounts for a specific expenditure register.
    def get_queryset(self):
        # queryset = super().get_queryset()
        queryset = RequestForPaymentOrInvoice.objects.filter(debit = self.kwargs['debit_pk'])
        return queryset

    # Add to context the data of the Account so that the user can see the register book it works on.
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        debit = Debit.objects.get( pk=self.kwargs['debit_pk'])
        context['debit'] = debit
        # invoices_query_set = RequestForPaymentOrInvoice.objects.filter(debit = self.kwargs['debit_pk'])
        invoices_query_set = context['object_list']
        invoice_list = [i.pk for i in list(invoices_query_set)]
        # payment_type_choice_list = [p.payment_type_choice for p in list(Payment.objects.filter(invoice__in = invoice_list))]
        payment_type_choice_list = [] # list(Payment.objects.filter(invoice__in = invoice_list).values_list('payment_type_choice'))[:][0]

        for i in invoice_list:
            try:
                invoice_payment_type_choice = Payment.objects.get(invoice = i).payment_type_choice
            except Payment.DoesNotExist:
                invoice_payment_type_choice = None
            payment_type_choice_list.append(invoice_payment_type_choice)

        context['data_list'] = zip(list(context['object_list']), payment_type_choice_list)

        # print(payment_status_list)
        s = Decimal(0.0)
        for q in invoices_query_set:
            s +=(q.invoice_charge_amount)
        context['debit_balance'] = debit.debit - s        
        return context


class RequestForPaymentOrInvoiceDeleteView(LoginRequiredMixin, DeleteView):
    model = RequestForPaymentOrInvoice

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit'] = Debit.objects.get(pk = self.kwargs['debit_pk'])
        
        try:
            invoice_payment = Payment.objects.get(invoice = self.kwargs['pk']).payment_amount
        except Payment.DoesNotExist:
            invoice_payment = Decimal(0.0)
        context['invoice_payment'] = invoice_payment
                
        if invoice_payment > 0:
            context['not_ok_to_delete'] = 1
        else:
            context['ok_to_delete'] = 1
        return context
  
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     credit = Credit.objects.get(pk = self.kwargs['pk'])
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     account.balance = account.balance - credit.credit*credit.disposed_percentage/Decimal(100.00)
    #     account.save()
    #     return super().form_valid(form)

    # template_name = 'expenditureregister_confirm_delete.html'
    def get_success_url(self) -> str:
        return reverse_lazy('invoice_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'], self.kwargs['debit_pk']))


class PaymentCreateView(LoginRequiredMixin, CreateView):
    model = Payment
    form_class = PaymentForm

    def get_initial(self):
        # account = get_object_or_404(Account, pk=self.kwargs['account_pk'])
        invoice = get_object_or_404(RequestForPaymentOrInvoice, pk=self.kwargs['invoice_pk'])
        return {'invoice': invoice, 'payment_amount': invoice.invoice_charge_amount} #'account': account, 

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit'] = Debit.objects.get( pk=self.kwargs['debit_pk'])
        context['invoice'] = RequestForPaymentOrInvoice.objects.get( pk=self.kwargs['invoice_pk'])
        invoice_payment = Payment.objects.filter(invoice = self.kwargs['invoice_pk']) 
        if invoice_payment.exists():
            payment_amount = invoice_payment[0].payment_amount # There should be one payment for each invoice.
        else:
            payment_amount = Decimal(0.00)

        if payment_amount > 0:
            context['not_ok_to_create'] = 1
        else:
            context['ok_to_create'] = 1
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     new_invoice = form.save(commit = False)
    #     debit = Debit.objects.get(pk = self.kwargs['debit_pk'])
    #     invoices_query_set = RequestForPaymentOrInvoice.objects.filter(debit = self.kwargs['debit_pk'])
    #     s = Decimal(0.0)
    #     for q in invoices_query_set:
    #         s +=(q.invoice_charge_amount)
    #     s += new_invoice.invoice_charge_amount
    #     print(s)
    #     return super().form_valid(form)

        # account.balance += new_credit.credit * new_credit.disposed_percentage/Decimal(100.00)
        # account.save()
        
            
    def get_success_url(self):
        return reverse_lazy('invoice_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'],self.kwargs['debit_pk']))


class PaymentListView(LoginRequiredMixin, ListView):
    model = Payment
    
    # Get the set of exit accounts for a specific expenditure register.
    def get_queryset(self):
        # queryset = super().get_queryset()
        queryset = Payment.objects.filter(invoice = self.kwargs['invoice_pk'])
        return queryset

    # Add to context the data of the Account so that the user can see the register book it works on.
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit']  = Debit.objects.get(pk = self.kwargs['debit_pk'])
        context['invoice'] = RequestForPaymentOrInvoice.objects.get(pk = self.kwargs['invoice_pk'])
        return context


class PaymentDetailView(LoginRequiredMixin, DetailView):
    model = Payment

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit'] = Debit.objects.get( pk=self.kwargs['debit_pk'])
        context['invoice'] = RequestForPaymentOrInvoice.objects.get( pk=self.kwargs['invoice_pk'])
        return context


class PaymentDeleteView(LoginRequiredMixin, DeleteView):
    model = Payment

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit'] = Debit.objects.get(pk = self.kwargs['debit_pk'])
        context['invoice'] = RequestForPaymentOrInvoice.objects.get(pk = self.kwargs['invoice_pk'])
        return context
  
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     credit = Credit.objects.get(pk = self.kwargs['pk'])
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     account.balance = account.balance - credit.credit*credit.disposed_percentage/Decimal(100.00)
    #     account.save()
    #     return super().form_valid(form)

    # template_name = 'expenditureregister_confirm_delete.html'
    def get_success_url(self) -> str:
        return reverse_lazy('payment_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'], self.kwargs['debit_pk'], self.kwargs['invoice_pk'],))


class PaymentUpdateView(LoginRequiredMixin, UpdateView):
    model = Payment
    # fields = '__all__' 
    # template_name = "requestforpaymentorinvoice_form.html"
    form_class = PaymentForm

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['register'] = ExpenditureRegister.objects.get(pk = self.kwargs['register_pk'])
        context['account'] = Account.objects.get(pk = self.kwargs['account_pk'])
        context['debit'] = Debit.objects.get( pk=self.kwargs['debit_pk'])
        context['invoice'] = RequestForPaymentOrInvoice.objects.get( pk=self.kwargs['invoice_pk'])
        return context
    
    # @transaction.atomic
    # def form_valid(self, form, **kwargs):
    #     old_credit = Credit.objects.get(pk = self.kwargs['pk'])
    #     new_credit = form.save(commit = False)
    #     account = Account.objects.get(pk = self.kwargs['account_pk'])
    #     account.balance =  account.balance + new_credit.credit * new_credit.disposed_percentage/Decimal(100.00) - old_credit.credit * old_credit.disposed_percentage/Decimal(100.00)
    #     account.save()
    #     form.save()
    #     return super().form_valid(form)

    def get_success_url(self) -> str:
        return reverse_lazy('payment_list', args = (self.kwargs['register_pk'], self.kwargs['account_pk'], self.kwargs['debit_pk'], self.kwargs['invoice_pk']))


def index(request):
    #return HttpResponse("Μητρώο Δεσμεύσεων")
    return render(request, 'expenditure_register/base.html')